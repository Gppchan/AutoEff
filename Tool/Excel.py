from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy


def create_table(worksheet: Worksheet, start_row: int, start_col: int, eff_list: list[float]):
    """
    在指定位置创建基础表格

    参数:
    worksheet: 工作表对象
    start_row: 起始行号
    start_col: 起始列号
    data: 表格中的数据
    """

    # 合并单元格
    # 第一列合并在一起, 第一行的第2到第5列合并, 第一行的第6到第9列合并
    worksheet.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=start_row + 2,
        end_column=start_col
    )
    worksheet.merge_cells(
        start_row=start_row,
        start_column=start_col + 1,
        end_row=start_row,
        end_column=start_col + 4
    )
    worksheet.merge_cells(
        start_row=start_row,
        start_column=start_col + 5,
        end_row=start_row,
        end_column=start_col + 8
    )

    # 添加第一行文本
    worksheet.cell(row=start_row, column=start_col).value = f""
    worksheet.cell(row=start_row, column=start_col + 1).value = "RAD"
    worksheet.cell(row=start_row, column=start_col + 5).value = "TOT"

    # 添加第二行文本
    worksheet.cell(row=start_row + 1, column=start_col + 1).value = "首"
    worksheet.cell(row=start_row + 1, column=start_col + 2).value = "峰值"
    worksheet.cell(row=start_row + 1, column=start_col + 3).value = "均值"
    worksheet.cell(row=start_row + 1, column=start_col + 4).value = "末"
    worksheet.cell(row=start_row + 1, column=start_col + 5).value = "首"
    worksheet.cell(row=start_row + 1, column=start_col + 6).value = "峰值"
    worksheet.cell(row=start_row + 1, column=start_col + 7).value = "均值"
    worksheet.cell(row=start_row + 1, column=start_col + 8).value = "末"

    # 添加第三行文本
    for i, _eff in enumerate(eff_list):
        worksheet.cell(row=start_row + 2, column=start_col + i + 1).value = _eff

    # 设置列宽，行高
    for _row_index in range(start_row, start_row + 3):
        worksheet.row_dimensions[_row_index].height = 20
    for _col_index in range(start_col, start_col + 9):
        col_letter = get_column_letter(_col_index)
        worksheet.column_dimensions[col_letter].width = 15

    # 设置单元格样式
    side = Side(style='thin')
    border = Border(left=side, right=side, top=side, bottom=side)
    alignment = Alignment(horizontal='center', vertical='center')
    for _row in worksheet.iter_rows(min_row=start_row, min_col=start_col, max_row=start_row + 2, max_col=start_col + 8):
        for __cell in _row:
            __cell.border = border
            __cell.alignment = alignment


def safe_copy_style(source_style):
    """安全复制样式，处理可能的布尔值和其他不可复制的对象"""
    if source_style is None:
        return None
    try:
        return copy.copy(source_style)
    except (TypeError, AttributeError):
        # 对于布尔值等不可复制的对象，直接返回值
        return source_style


def move_table(worksheet, source_info, row_offset=0, col_offset=0, copy=False):
    """
    移动或复制表格到新位置

    参数:
    worksheet: 工作表对象
    source_info: 源表格信息（包含起始行、列和结束行、列）
    row_offset: 行偏移量（正数向下，负数向上）
    col_offset: 列偏移量（正数向右，负数向左）
    copy: 是否复制（True为复制，False为移动）
    """
    # 计算目标位置
    target_start_row = source_info['start_row'] + row_offset
    target_start_col = source_info['start_col'] + col_offset
    target_end_row = source_info['end_row'] + row_offset
    target_end_col = source_info['end_col'] + col_offset

    # 先复制合并单元格信息
    merged_cells_to_copy = []
    for merged_range in list(worksheet.merged_cells.ranges):
        if (merged_range.min_row >= source_info['start_row'] and
                merged_range.max_row <= source_info['end_row'] and
                merged_range.min_col >= source_info['start_col'] and
                merged_range.max_col <= source_info['end_col']):
            # 计算新的合并范围
            new_min_row = merged_range.min_row + row_offset
            new_max_row = merged_range.max_row + row_offset
            new_min_col = merged_range.min_col + col_offset
            new_max_col = merged_range.max_col + col_offset

            merged_cells_to_copy.append({
                'original': merged_range,
                'new_min_row': new_min_row,
                'new_max_row': new_max_row,
                'new_min_col': new_min_col,
                'new_max_col': new_max_col
            })

    # 复制行高
    for row in range(source_info['start_row'], source_info['end_row'] + 1):
        target_row = row + row_offset
        if row in worksheet.row_dimensions:
            worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[row].height

    # 复制列宽
    for col in range(source_info['start_col'], source_info['end_col'] + 1):
        source_col_letter = get_column_letter(col)
        target_col = col + col_offset
        target_col_letter = get_column_letter(target_col)
        if source_col_letter in worksheet.column_dimensions:
            worksheet.column_dimensions[target_col_letter].width = worksheet.column_dimensions[source_col_letter].width

    # 复制单元格内容和样式
    for row in range(source_info['start_row'], source_info['end_row'] + 1):
        for col in range(source_info['start_col'], source_info['end_col'] + 1):
            source_cell = worksheet.cell(row=row, column=col)
            target_cell = worksheet.cell(row=row + row_offset, column=col + col_offset)

            # 复制值
            target_cell.value = source_cell.value

            # 安全复制样式
            if source_cell.has_style:
                target_cell.font = safe_copy_style(source_cell.font)
                target_cell.border = safe_copy_style(source_cell.border)
                target_cell.fill = safe_copy_style(source_cell.fill)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = safe_copy_style(source_cell.protection)
                target_cell.alignment = safe_copy_style(source_cell.alignment)

    # 创建新的合并单元格
    for merged_info in merged_cells_to_copy:
        worksheet.merge_cells(
            start_row=merged_info['new_min_row'],
            start_column=merged_info['new_min_col'],
            end_row=merged_info['new_max_row'],
            end_column=merged_info['new_max_col']
        )

    # 如果不是复制，则清除源表格内容
    if not copy:
        for row in range(source_info['start_row'], source_info['end_row'] + 1):
            for col in range(source_info['start_col'], source_info['end_col'] + 1):
                worksheet.cell(row=row, column=col).value = None

        # 取消源合并单元格
        for merged_info in merged_cells_to_copy:
            worksheet.unmerge_cells(str(merged_info['original']))

    return {
        'start_row': target_start_row,
        'start_col': target_start_col,
        'end_row': target_end_row,
        'end_col': target_end_col
    }


# 创建一个新的工作簿
wb = Workbook()
ws = wb.active
ws.title = "表格操作示例"

# 创建第一个基础表格
# table1_info = create_base_table(ws, 1, 1, 1)
#
# # 复制表格到下方（行偏移4，列偏移0）
# table2_info = move_table(ws, table1_info, 4, 0, copy=True)
#
# # 再次复制表格到右侧（行偏移0，列偏移10）
# table3_info = move_table(ws, table2_info, 0, 10, copy=True)
#
# # 保存文件
# wb.save("表格操作示例.xlsx")
# print("Excel文件已创建成功！")
