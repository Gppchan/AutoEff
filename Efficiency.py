from pathlib import Path
from datetime import datetime

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
import numpy as np

from Frequency import Frequency, FrequencyBand, FrequencyManager
import CST

def cal_antenna_eff_map(cst: CST.Contents, antenna_table: list[tuple[str, str, str]]) -> dict[
    str, dict[str, tuple[float]]]:
    # 提取效率图
    eff_graphs: list[CST.Graph] = CST.Graph.extract_graph(cst.model_res, "System Tot. Efficiency")
    eff_charts: list[CST.LineChart] = []
    for _graph in eff_graphs:
        line_chart = CST.LineChart.parse_sig_file(cst.result_dir, _graph)
        eff_charts.append(line_chart)

    antenna_eff_map: dict[str, dict[str, tuple[float]]] = {}
    for _antenna_name, _graph_name, _freq_text in antenna_table:
        if not _freq_text:
            continue

        eff_chart: CST.LineChart
        for __chart in eff_charts:
            if __chart.graph.name == _graph_name:
                eff_chart = __chart
                break

        freq_eff_map: dict[str, tuple[float]] = {}
        freq_texts: list[str] = [item.strip() for item in _freq_text.split(",")]
        for __freq_text in freq_texts:
            if not  __freq_text:
                continue
            freq = FrequencyManager.parse_freq_text(__freq_text)[0]
            freq_inf: float = 0
            freq_sup: float = 0
            if type(freq) == Frequency:
                freq_inf = freq.inf / 1000.0
                freq_sup = freq.sup / 1000.0
            elif type(freq) == FrequencyBand:
                if freq.is_duplex():
                    freq_inf = freq.downlink.inf / 1000.0
                    freq_sup = freq.downlink.sup / 1000.0
                else:
                    freq_inf = freq.link.inf / 1000.0
                    freq_sup = freq.link.sup / 1000.0

            is_valid_inf: bool = eff_chart.x.min() <= freq_inf < eff_chart.x.max()
            is_valid_sup: bool = eff_chart.x.min() < freq_sup <= eff_chart.x.max()
            if not is_valid_inf or not is_valid_sup:
                freq_eff_map |= {__freq_text: (-1, -1, -1, -1)}
                continue

            inf_index = np.searchsorted(eff_chart.x, freq_inf)
            sup_index = np.searchsorted(eff_chart.x, freq_sup)
            if freq_inf != eff_chart.x[inf_index]:
                inf_index = inf_index - 1

            sub_eff = np.real(eff_chart.y[inf_index: sup_index + 1])
            first_eff = float(sub_eff[0])
            max_eff = float(sub_eff.max())
            last_eff = float(sub_eff[-1])
            avg_eff = float(sub_eff.mean())
            freq_eff_map |= {freq.name: (first_eff, max_eff, last_eff, avg_eff)}

        antenna_eff_map |= {_antenna_name: freq_eff_map}
    return antenna_eff_map

def export_antenna_eff_map(antenna_eff_map: dict[str, dict[str, tuple[float]]], excel_dir : Path):
    wb = Workbook()
    ws : Worksheet = wb.active
    ws.title = "效率"
    font = Font(bold=True, italic=False, name="微软雅黑", size=14)
    alignment = Alignment(horizontal='center', vertical='center')
    for i, _antenna in enumerate(antenna_eff_map):
        row = i * 3 + 1
        eff_map = antenna_eff_map[_antenna]
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=1)
        ws.cell(row=row, column=1).value = _antenna
        ws.cell(row=row, column=1).alignment = alignment
        ws.cell(row=row, column=1).font = font
        for __row_index in range(row, row + 3):
            ws.row_dimensions[__row_index].height = 20
        for j, __freq  in enumerate(eff_map):
            col = j + 2
            effs = eff_map[__freq]
            ws.cell(row=row, column=col).value = __freq
            ws.cell(row=row, column=col).alignment = alignment
            ws.cell(row=row, column=col).font = font
            ws.cell(row=row+1, column=col).value = f"{effs[0]:.2f}_{effs[1]:.2f}_{effs[2]:.2f}"
            ws.cell(row=row+1, column=col).alignment = alignment
            ws.cell(row=row+2, column=col).value = f"{effs[3]:.2f}"
            ws.cell(row=row+2, column=col).alignment = alignment
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 25

    # 保存文件
    wb.save(excel_dir / f"eff_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx")
